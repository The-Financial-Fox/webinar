import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from prophet import Prophet
from fpdf import FPDF
import openpyxl
from io import BytesIO
from dotenv import load_dotenv
import os

# Load environment variables (optional for Groq API integration in the future)
load_dotenv()
GROQ_API_KEY = os.getenv("GROQ_API_KEY")  # Not used yet, but included for future AI insights

# Dummy data for demonstration
revenue_data = pd.DataFrame({
    'Month': pd.date_range(start='2023-01-01', periods=12, freq='M'),
    'Revenue ($M)': [5.2, 5.5, 6.0, 5.8, 6.2, 6.5, 6.8, 7.0, 7.2, 7.5, 7.8, 8.0]
})

expenses_data = pd.DataFrame({
    'Category': ['Marketing', 'Operations', 'R&D'],
    'Expenses ($M)': [1.5, 2.0, 1.0]
})

# Streamlit app configuration
st.set_page_config(page_title="The Financial Fox FP&A Tool", layout="wide")

# Custom CSS for The Financial Fox branding (gold: #FFD700, navy blue: #1E3A8A)
st.markdown("""
    <style>
    .stApp {
        background-color: #1E3A8A;
        color: #FFFFFF;
    }
    .stButton>button {
        background-color: #FFD700;
        color: #1E3A8A;
    }
    .stTextInput>div>input {
        background-color: #FFFFFF;
        color: #1E3A8A;
    }
    .stSlider>div>div {
        background-color: #FFD700;
    }
    </style>
""", unsafe_allow_html=True)

# Sidebar navigation
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Dashboard", "Scenario Analysis", "Forecasting", "Budgeting", "Report Generation"])

# Initialize session state for data persistence
if 'budget_data' not in st.session_state:
    st.session_state.budget_data = expenses_data.copy()
if 'forecast_data' not in st.session_state:
    st.session_state.forecast_data = None

# Dashboard Page
if page == "Dashboard":
    st.title("📊 The Financial Fox FP&A Tool")
    st.write("Welcome to your financial planning and analysis assistant!")
    st.image("https://via.placeholder.com/150x50.png?text=The+Financial+Fox+Logo", use_column_width=True)

    # KPIs
    total_revenue = revenue_data['Revenue ($M)'].sum()
    profit_margin = (total_revenue - expenses_data['Expenses ($M)'].sum()) / total_revenue * 100
    cash_flow = total_revenue - expenses_data['Expenses ($M)'].sum()

    col1, col2, col3 = st.columns(3)
    col1.metric("Total Revenue", f"${total_revenue:.2f}M")
    col2.metric("Profit Margin", f"{profit_margin:.2f}%")
    col3.metric("Cash Flow", f"${cash_flow:.2f}M")

    # Revenue Trend Chart
    fig = px.line(revenue_data, x='Month', y='Revenue ($M)', title='Revenue Over Time')
    fig.update_layout(template="plotly_white")
    st.plotly_chart(fig)

# Scenario Analysis Page
elif page == "Scenario Analysis":
    st.header("🔍 Scenario Analysis")
    st.write("Adjust variables to see how they impact forecasts and budgets.")

    # Sliders for scenario variables
    revenue_growth = st.slider("Revenue Growth Rate (%)", -20.0, 20.0, 0.0, step=0.5)
    cost_reduction = st.slider("Cost Reduction (%)", 0.0, 20.0, 0.0, step=0.5)

    # Adjust revenue and expenses based on scenarios
    adjusted_revenue = revenue_data['Revenue ($M)'] * (1 + revenue_growth / 100)
    adjusted_expenses = st.session_state.budget_data['Expenses ($M)'] * (1 - cost_reduction / 100)

    # Calculate adjusted KPIs
    adjusted_total_revenue = adjusted_revenue.sum()
    adjusted_profit_margin = (adjusted_total_revenue - adjusted_expenses.sum()) / adjusted_total_revenue * 100
    adjusted_cash_flow = adjusted_total_revenue - adjusted_expenses.sum()

    # Display adjusted KPIs
    col1, col2, col3 = st.columns(3)
    col1.metric("Adjusted Total Revenue", f"${adjusted_total_revenue:.2f}M")
    col2.metric("Adjusted Profit Margin", f"{adjusted_profit_margin:.2f}%")
    col3.metric("Adjusted Cash Flow", f"${adjusted_cash_flow:.2f}M")

    # Adjusted Revenue Chart
    adjusted_revenue_df = pd.DataFrame({'Month': revenue_data['Month'], 'Adjusted Revenue ($M)': adjusted_revenue})
    fig = px.line(adjusted_revenue_df, x='Month', y='Adjusted Revenue ($M)', title='Adjusted Revenue Over Time')
    fig.update_layout(template="plotly_white")
    st.plotly_chart(fig)

# Forecasting Page
elif page == "Forecasting":
    st.header("🔮 Forecasting")
    st.write("Select a forecasting model and generate predictions.")

    model_type = st.selectbox("Select Forecasting Model", ["ARIMA", "Prophet", "Exponential Smoothing"])
    forecast_horizon = st.number_input("Forecast Horizon (months)", min_value=1, value=3)

    if st.button("Generate Forecast"):
        with st.spinner("Generating forecast..."):
            if model_type == "ARIMA":
                model = ARIMA(revenue_data['Revenue ($M)'], order=(1,1,1))
                fitted_model = model.fit()
                forecast = fitted_model.forecast(steps=forecast_horizon)
            elif model_type == "Prophet":
                prophet_df = revenue_data.rename(columns={'Month': 'ds', 'Revenue ($M)': 'y'})
                model = Prophet()
                model.fit(prophet_df)
                future = model.make_future_dataframe(periods=forecast_horizon, freq='M')
                forecast = model.predict(future)['yhat'].tail(forecast_horizon).values
            elif model_type == "Exponential Smoothing":
                model = ExponentialSmoothing(revenue_data['Revenue ($M)'], trend='add', seasonal='add', seasonal_periods=12)
                fitted_model = model.fit()
                forecast = fitted_model.forecast(steps=forecast_horizon)

            forecast_dates = pd.date_range(start=revenue_data['Month'].iloc[-1] + pd.DateOffset(months=1), periods=forecast_horizon, freq='M')
            forecast_df = pd.DataFrame({'Month': forecast_dates, 'Forecast ($M)': forecast})
            st.session_state.forecast_data = forecast_df

            fig = go.Figure()
            fig.add_trace(go.Scatter(x=revenue_data['Month'], y=revenue_data['Revenue ($M)'], mode='lines', name='Historical'))
            fig.add_trace(go.Scatter(x=forecast_df['Month'], y=forecast_df['Forecast ($M)'], mode='lines', name='Forecast'))
            fig.update_layout(title=f"{model_type} Forecast", xaxis_title="Month", yaxis_title="Revenue ($M)", template="plotly_white")
            st.plotly_chart(fig)

# Budgeting Page
elif page == "Budgeting":
    st.header("💰 Budgeting")
    st.write("Manage your budgets and compare with actuals.")

    budget_version = st.selectbox("Select Budget Version", ["Base", "Optimistic", "Pessimistic"])
    edited_df = st.data_editor(st.session_state.budget_data, num_rows="dynamic")

    if st.button("Save Budget"):
        st.session_state.budget_data = edited_df
        st.success(f"Budget '{budget_version}' saved successfully!")

    # Calculate variance (assuming 'Actual ($M)' could be added later)
    edited_df['Variance ($M)'] = edited_df['Expenses ($M)'] - edited_df.get('Actual ($M)', edited_df['Expenses ($M)'])
    st.write("Budget vs Actuals with Variance:")
    st.dataframe(edited_df)

# Report Generation Page
elif page == "Report Generation":
    st.header("📄 Report Generation")
    st.write("Select components to include in your report.")

    include_dashboard = st.checkbox("Include Dashboard KPIs")
    include_forecast = st.checkbox("Include Forecast Chart")
    include_budget = st.checkbox("Include Budget Table")

    if st.button("Generate PDF Report"):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="The Financial Fox FP&A Report", ln=True, align='C')

        if include_dashboard:
            pdf.cell(200, 10, txt="Dashboard KPIs", ln=True)
            pdf.cell(200, 10, txt=f"Total Revenue: ${total_revenue:.2f}M", ln=True)
            pdf.cell(200, 10, txt=f"Profit Margin: {profit_margin:.2f}%", ln=True)
            pdf.cell(200, 10, txt=f"Cash Flow: ${cash_flow:.2f}M", ln=True)

        if include_forecast and st.session_state.forecast_data is not None:
            pdf.cell(200, 10, txt="Forecast Data", ln=True)
            for index, row in st.session_state.forecast_data.iterrows():
                pdf.cell(200, 10, txt=f"{row['Month'].strftime('%Y-%m')}: ${row['Forecast ($M)']:.2f}M", ln=True)

        if include_budget:
            pdf.cell(200, 10, txt="Budget Data", ln=True)
            for index, row in st.session_state.budget_data.iterrows():
                pdf.cell(200, 10, txt=f"{row['Category']}: Budget ${row['Expenses ($M)']:.2f}M, Actual ${row.get('Actual ($M)', 0):.2f}M", ln=True)

        pdf_output = BytesIO()
        pdf.output(pdf_output)
        st.download_button(label="Download PDF Report", data=pdf_output.getvalue(), file_name="fpna_report.pdf", mime="application/pdf")

# Placeholder for future AI insights
st.sidebar.markdown("---")
st.sidebar.write("AI Insights: Coming soon with Groq API integration for proactive analysis.")
